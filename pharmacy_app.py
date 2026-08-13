import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
from datetime import datetime, timedelta
from pharmacy_db import update_database, backup_db, CURRENCY, CURRENCY_SYMBOL
from pharmacy_backend import DB
from pharmacy_pos import POSFrame, is_expired_or_near

def safe_get_id(tree_selection, tree_object):
    """استخراج المعرّف الرقمي النقي من العمود الأول للسطر المختار"""
    if not tree_selection:
        return None
    try:
        # tree_selection قد يكون tuple
        item_id = tree_selection[0] if isinstance(tree_selection, (tuple, list)) else tree_selection
        vals = tree_object.item(item_id, "values")
        if vals and len(vals) > 0:
            return int(vals[0])
        return None
    except Exception:
        return None

class PharmacyApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("نظام إدارة الصيدلية - Pharmacy Clinical Professional")
        self.geometry("1250x850")
        update_database()
        self.db = DB()
        self.current_user_role = "cashier"

        self._set_modern_theme()
        self.withdraw()
        self._show_license_window()

    def _set_modern_theme(self):
        """إعداد وتطبيق الألوان الاحترافية العصرية للنظام بالكامل"""
        self.style = ttk.Style()
        self.style.theme_use("clam")

        bg_primary = "#f4f7f6"
        teal_dark = "#115e59"
        text_dark = "#1e293b"

        self.configure(bg=bg_primary)

        self.style.configure(".", background=bg_primary, foreground=text_dark, font=("Segoe UI", 10))
        self.style.configure("TNotebook", background=bg_primary, padding=4)
        self.style.configure("TNotebook.Tab", background="#e2e8f0", foreground=text_dark, font=("Segoe UI", 10, "bold"), padding=(15, 6))
        self.style.map("TNotebook.Tab", background=[("selected", teal_dark)], foreground=[("selected", "#ffffff")])
        self.style.configure("TButton", background="#e2e8f0", foreground=text_dark, font=("Segoe UI", 10, "bold"), padding=5)
        self.style.map("TButton", background=[("active", "#14b8a6"), ("pressed", "#115e59")], foreground=[("active", "#ffffff")])
        self.style.configure("Treeview", background="#ffffff", foreground=text_dark, rowheight=26, fieldbackground="#ffffff")
        self.style.configure("Treeview.Heading", background="#cbd5e1", foreground=text_dark, font=("Segoe UI", 10, "bold"))
        self.style.map("Treeview", background=[("selected", "#115e59")], foreground=[("selected", "#ffffff")])

    def _show_license_window(self):
        """نافذة التفعيل - تظهر قبل تسجيل الدخول"""
        is_valid, msg = self.db.check_license()
        if is_valid:
            # التفعيل صالح، انتقل مباشرة لتسجيل الدخول
            self._show_login_window()
            return

        self.lic_win = tk.Toplevel(self)
        self.lic_win.title("تفعيل البرنامج / Activation")
        self.lic_win.geometry("420x280")
        self.lic_win.resizable(False, False)
        self.lic_win.protocol("WM_DELETE_WINDOW", self.on_close)
        self.lic_win.configure(bg="#f4f7f6")

        ttk.Label(self.lic_win, text="🔐 تفعيل البرنامج", font=("Segoe UI", 16, "bold"), foreground="#115e59").pack(pady=(15, 5))
        ttk.Label(self.lic_win, text="البرنامج يتطلب تفعيل سنوي", font=("Segoe UI", 10), foreground="#dc2626").pack()

        ttk.Label(self.lic_win, text=msg, font=("Segoe UI", 9), foreground="#64748b", wraplength=350).pack(pady=10)

        ttk.Label(self.lic_win, text="أدخل كود التفعيل:", font=("Segoe UI", 11, "bold")).pack(pady=(10, 2))
        self.lic_code_var = tk.StringVar()
        ttk.Entry(self.lic_win, textvariable=self.lic_code_var, width=35, show="*").pack()

        ttk.Button(self.lic_win, text="تفعيل", command=self._process_license, width=20).pack(pady=15)
        ttk.Label(self.lic_win, text="للحصول على كود التفعيل، تواصل مع الدعم", font=("Segoe UI", 8), foreground="#94a3b8").pack()

        self.lic_win.grab_set()
        self.lic_win.focus_force()

    def _process_license(self):
        code = self.lic_code_var.get().strip()
        if not code:
            messagebox.showwarning("تنبيه", "أدخل كود التفعيل", parent=self.lic_win)
            return
        ok, msg = self.db.activate_license(code)
        if ok:
            messagebox.showinfo("تم", msg, parent=self.lic_win)
            self.lic_win.destroy()
            self._show_login_window()
        else:
            messagebox.showerror("خطأ", msg, parent=self.lic_win)

    def _show_login_window(self):
        self.login_win = tk.Toplevel(self)
        self.login_win.title("تسجيل الدخول / Login")
        self.login_win.geometry("380x260")
        self.login_win.resizable(False, False)
        self.login_win.protocol("WM_DELETE_WINDOW", self.on_close)
        self.login_win.configure(bg="#f4f7f6")

        ttk.Label(self.login_win, text="🔐 تسجيل الدخول", font=("Segoe UI", 14, "bold"), foreground="#115e59").pack(pady=(20, 10))

        ttk.Label(self.login_win, text="اسم المستخدم:", font=("Segoe UI", 10)).pack(pady=(10, 2))
        self.user_var = tk.StringVar()
        ttk.Entry(self.login_win, textvariable=self.user_var, width=32).pack()

        ttk.Label(self.login_win, text="كلمة المرور:", font=("Segoe UI", 10)).pack(pady=(10, 2))
        self.pass_var = tk.StringVar()
        ttk.Entry(self.login_win, textvariable=self.pass_var, width=32, show="*").pack()

        ttk.Button(self.login_win, text="دخول", command=self._process_login, width=20).pack(pady=20)
        ttk.Label(self.login_win, text="افتراضي: admin / admin123", font=("Segoe UI", 8), foreground="#94a3b8").pack()
        self.login_win.grab_set()
        self.login_win.focus_force()

    def _process_login(self):
        u = self.user_var.get().strip()
        p = self.pass_var.get().strip()
        user_row = self.db.check_login(u, p)

        if user_row:
            self.current_user_role = user_row["role"]
            self.current_user_name = user_row["username"]
            self.login_win.destroy()
            self.deiconify()
            self._build_ui()
            self._apply_role_permissions()
            self._reload_all_tabs()
            self.notebook.bind("<<NotebookTabChanged>>", lambda e: self._on_tab_changed())
            self.after(800, self._check_low_stock_popup)
        else:
            messagebox.showerror("خطأ", "بيانات الدخول غير صحيحة!", parent=self.login_win)

    def _apply_role_permissions(self):
        if self.current_user_role == "cashier":
            for tab_name in ["المخزن", "المسترجعات", "التقارير", "الأطباء", "المرضى", "المستخدمين", "التنبيهات"]:
                for idx in range(self.notebook.index("end")):
                    if tab_name in self.notebook.tab(idx, "text"):
                        self.notebook.hide(idx)
                        break
            messagebox.showinfo("مرحباً بك", f"تم تسجيل الدخول بصلاحية: كاشير\nالمستخدم: {self.current_user_name}")
        else:
            messagebox.showinfo("مرحباً بك", f"تم تسجيل الدخول بصلاحية: مدير النظام\nالمستخدم: {self.current_user_name}")

    def _on_tab_changed(self):
        tab_idx = self.notebook.index(self.notebook.select())
        tab_text = self.notebook.tab(tab_idx, "text")
        if "المخزن" in tab_text:
            self._stock_reload()
        elif "المسترجعات" in tab_text:
            self._returns_reload()
        elif "التقارير" in tab_text:
            self._reports_reload_default()
        elif "الأطباء" in tab_text:
            self._doctors_reload()
        elif "المرضى" in tab_text:
            self._patients_reload()
        elif "المستخدمين" in tab_text:
            self._users_reload()
        elif "التنبيهات" in tab_text:
            self._check_alerts()

    def _check_low_stock_popup(self):
        if self.current_user_role != "admin":
            return
        low_meds = self.db.get_low_stock_medicines()
        if low_meds:
            msg = "⚠️ تنبيه: الأصناف التالية وصلت للحد الأدنى أو نفدت:\n\n"
            for m in low_meds[:15]:
                msg += f"- {m['name']} (المتبقي: {m['quantity']} | الحد: {m['min_limit']})\n"
            messagebox.showwarning("تنبيه المخزون المنخفض", msg, parent=self)

    def _build_ui(self):
        self.notebook = ttk.Notebook(self)
        self.notebook.pack(fill="both", expand=True, padx=8, pady=8)

        self.pos_tab = POSFrame(self.notebook, self)
        self.notebook.add(self.pos_tab, text="🛒 نقطة البيع / POS")

        self.frame_stock = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_stock, text="📦 المخزن / Stock")
        self._build_stock(self.frame_stock)

        self.frame_returns = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_returns, text="↩️ المسترجعات / Returns")
        self._build_returns(self.frame_returns)

        self.frame_reports = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_reports, text="📊 التقارير / Reports")
        self._build_reports(self.frame_reports)

        self.frame_doctors = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_doctors, text="👨‍⚕️ الأطباء / Doctors")
        self._build_doctors_tab(self.frame_doctors)

        self.frame_patients = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_patients, text="🧑‍⚕️ المرضى / Patients")
        self._build_patients_tab(self.frame_patients)

        self.frame_users = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_users, text="👤 المستخدمين / Users")
        self._build_users_tab(self.frame_users)

        self.frame_alerts = ttk.Frame(self.notebook)
        self.notebook.add(self.frame_alerts, text="⚠️ التنبيهات / Alerts")
        self._build_alerts(self.frame_alerts)

        self.status_var = tk.StringVar(value=f"جاهز | المستخدم: {getattr(self, 'current_user_name', 'unknown')}")
        self.statusbar = ttk.Label(self, textvariable=self.status_var, anchor="w", relief="sunken")
        self.statusbar.pack(fill="x", side="bottom")

    def _reload_all_tabs(self):
        self._stock_reload()
        self._returns_reload()
        self._reports_reload_default()
        self._users_reload()
        self._doctors_reload()
        self._patients_reload()
        self._check_alerts()
        try:
            self.pos_tab.refresh_clinical_combos()
        except:
            pass

    # ==================== المخزن ====================
    def _build_stock(self, parent):
        search_frame = ttk.Frame(parent)
        search_frame.pack(fill="x", padx=8, pady=6)
        ttk.Label(search_frame, text="🔍 بحث:").pack(side="left")
        self.stock_search_var = tk.StringVar()
        se = ttk.Entry(search_frame, textvariable=self.stock_search_var, width=30)
        se.pack(side="left", padx=6)
        se.bind("<KeyRelease>", lambda e: self._stock_reload())
        ttk.Button(search_frame, text="🔄 تحديث", command=self._stock_reload).pack(side="left", padx=4)
        ttk.Button(search_frame, text="🗑️ حذف دواء", command=self._stock_delete).pack(side="left", padx=4)

        left = ttk.Frame(parent)
        left.pack(side="left", fill="y", padx=8, pady=8)

        self.vars = {k: tk.StringVar() for k in ["bar", "name", "sci", "qty", "min", "exp", "pur", "sell"]}
        for k, txt in [("bar","Barcode *"), ("name","Name *"), ("sci","Scientific name"), ("qty","Quantity"), ("min","Min limit"), ("exp","Expiry (YYYY-MM-DD)"), ("pur","Purchase price"), ("sell","Selling price")]:
            ttk.Label(left, text=txt).pack(anchor="w")
            ttk.Entry(left, textvariable=self.vars[k]).pack(fill="x")

        ttk.Button(left, text="💾 حفظ / إضافة دواء", command=self._stock_save).pack(fill="x", pady=6)
        ttk.Button(left, text="🧹 مسح الحقول", command=self._stock_clear_fields).pack(fill="x", pady=2)

        right = ttk.Frame(parent)
        right.pack(side="right", fill="both", expand=True, padx=8, pady=8)

        cols = ("id", "barcode", "name", "scientific", "qty", "status", "expiry")
        self.tree_stock = ttk.Treeview(right, columns=cols, show="headings")
        for c, t in zip(cols, ["#", "باركود", "الاسم", "الاسم العلمي", "الكمية", "الحالة", "الصلاحية"]):
            self.tree_stock.heading(c, text=t)
        self.tree_stock.column("id", width=40)
        self.tree_stock.column("barcode", width=100)
        self.tree_stock.column("name", width=180)
        self.tree_stock.column("scientific", width=130)
        self.tree_stock.column("qty", width=60, anchor="center")
        self.tree_stock.column("status", width=70, anchor="center")
        self.tree_stock.column("expiry", width=100, anchor="center")
        self.tree_stock.pack(fill="both", expand=True)
        self.tree_stock.bind("<Double-1>", self._stock_load_to_entries)
        self.tree_stock.tag_configure("low_stock", foreground="#dc2626", font=("Segoe UI", 9, "bold"))
        self.tree_stock.tag_configure("near_expiry", foreground="#ea580c")
        self.tree_stock.tag_configure("expired", foreground="#dc2626", font=("Segoe UI", 9, "bold"))

    def _stock_clear_fields(self):
        for v in self.vars.values():
            v.set("")

    def _stock_load_to_entries(self, event=None):
        sel = self.tree_stock.selection()
        if not sel:
            return
        iid = sel[0]
        med = self.db.get_medicine_by_id(int(iid))
        if not med:
            return
        self.vars["bar"].set(med["barcode"] or "")
        self.vars["name"].set(med["name"] or "")
        self.vars["sci"].set(med["scientific_name"] or "")
        self.vars["qty"].set(str(med["quantity"] or 0))
        self.vars["min"].set(str(med["min_limit"] or 0))
        self.vars["exp"].set(med["expiry_date"] or "")
        self.vars["pur"].set(str(med["purchase_price"] or 0.0))
        self.vars["sell"].set(str(med["selling_price"] or 0.0))

    def _stock_save(self):
        b = self.vars["bar"].get().strip()
        n = self.vars["name"].get().strip()
        s = self.vars["sci"].get().strip()
        if not b or not n:
            messagebox.showwarning("مطلوب", "الباركود والاسم مطلوبان")
            return
        existing = self.db.get_medicine_by_barcode(b)
        if existing and existing["name"] != n:
            if not messagebox.askyesno("تأكيد", f"الباركود مسجل باسم ({existing['name']})\nهل تريد استبداله؟"):
                return
        try:
            q = max(0, int(self.vars["qty"].get() or 0))
            m = max(0, int(self.vars["min"].get() or 0))
        except:
            messagebox.showerror("خطأ", "الكمية والحد الأدنى يجب أن يكونا أرقاماً صحيحة")
            return
        exp = self.vars["exp"].get().strip() or None
        if exp:
            try:
                datetime.strptime(exp, "%Y-%m-%d")
            except:
                messagebox.showerror("خطأ", "صيغة التاريخ يجب أن تكون YYYY-MM-DD")
                return
        try:
            p = max(0.0, float(self.vars["pur"].get() or 0.0))
            sl = max(0.0, float(self.vars["sell"].get() or 0.0))
        except:
            messagebox.showerror("خطأ", "الأسعار يجب أن تكون أرقاماً")
            return
        try:
            self.db.add_or_update_medicine(b, n, s, q, m, exp, p, sl)
            messagebox.showinfo("تم", "تم الحفظ بنجاح")
            self._stock_reload()
            self._stock_clear_fields()
            self._check_alerts()
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل الحفظ: {e}")

    def _stock_reload(self):
        for i in self.tree_stock.get_children():
            self.tree_stock.delete(i)
        search = self.stock_search_var.get().strip()
        for r in self.db.list_medicines():
            if search and search.lower() not in (r["name"] or "").lower() and search not in (r["barcode"] or "") and search.lower() not in (r["scientific_name"] or "").lower():
                continue
            qty = r["quantity"] or 0
            minl = r["min_limit"] or 0
            if qty <= 0:
                status = "❌ نفذ"
                tag = "low_stock"
            elif qty <= minl and minl > 0:
                status = "⚠️ منخفض"
                tag = "low_stock"
            else:
                status = "✅ OK"
                tag = ""

            exp = r["expiry_date"] or ""
            expired, near = is_expired_or_near(exp, days=30)
            if expired:
                exp_display = f"🔴 {exp}"
                tag = "expired"
            elif near:
                exp_display = f"🟡 {exp}"
                if not tag:
                    tag = "near_expiry"
            else:
                exp_display = f"🟢 {exp}" if exp else ""

            iid = self.tree_stock.insert("", tk.END, iid=str(r["id"]),
                values=(r["id"], r["barcode"] or "", r["name"] or "", r["scientific_name"] or "غير محدد", qty, status, exp_display))
            if tag:
                self.tree_stock.item(iid, tags=(tag,))

    def _stock_delete(self):
        sel = self.tree_stock.selection()
        if not sel:
            messagebox.showwarning("اختر", "اختر دواءً للحذف")
            return
        iid = sel[0]
        if not messagebox.askyesno("تأكيد", "هل أنت متأكد من حذف هذا الدواء؟"):
            return
        ok, msg = self.db.delete_medicine(int(iid))
        if ok:
            messagebox.showinfo("تم", msg)
            self._stock_reload()
            self._check_alerts()
        else:
            messagebox.showerror("خطأ", msg)

    # ==================== المسترجعات ====================
    def _build_returns(self, parent):
        top = ttk.Frame(parent)
        top.pack(fill="x", padx=8, pady=6)

        ttk.Label(top, text="🔍 بحث برقم الفاتورة:").pack(side="left")
        self.return_search_var = tk.StringVar()
        se = ttk.Entry(top, textvariable=self.return_search_var, width=22)
        se.pack(side="left", padx=4)
        se.bind("<KeyRelease>", self._return_search_key)

        ttk.Button(top, text="📋 استعراض الفاتورة", command=self._preview_invoice_content).pack(side="left", padx=2)
        ttk.Button(top, text="↩️ استرجاع كامل", command=self._full_return).pack(side="left", padx=2)
        ttk.Button(top, text="↩️ استرجاع جزئي", command=self._partial_return).pack(side="left", padx=2)

        self.lb_return_suggest = tk.Listbox(parent, height=3, font=("TkDefaultFont", 10))
        self.lb_return_suggest.pack(fill="x", padx=8)
        self.lb_return_suggest.bind("<<ListboxSelect>>", self._return_suggest_select)

        lbl_p = ttk.Label(parent, text="↓↓ محتويات الفاتورة قيد المعاينة ↓↓", font=("Segoe UI", 9, "bold"))
        lbl_p.pack(anchor="w", padx=8)

        self.tree_invoice_preview = ttk.Treeview(parent, columns=("barcode", "name", "qty", "price", "total"), show="headings", height=4)
        for c, t in zip(("barcode", "name", "qty", "price", "total"), ["باركود", "الدواء", "الكمية", "السعر", "الإجمالي"]):
            self.tree_invoice_preview.heading(c, text=t)
        self.tree_invoice_preview.pack(fill="x", padx=8, pady=2)

        ttk.Label(parent, text="سجل العمليات المرتجعة السابقة (نقرتين للتفاصيل):").pack(anchor="w", padx=8)
        self.tree_returns = ttk.Treeview(parent, columns=("id", "num", "date", "invoice", "total", "by"), show="headings", height=5)
        for c, t in zip(["id", "num", "date", "invoice", "total", "by"], ["#", "رقم الإرجاع", "التاريخ", "الفاتورة الأصلية", "المسترد", "المستخدم"]):
            self.tree_returns.heading(c, text=t)
        self.tree_returns.column("id", width=40)
        self.tree_returns.column("num", width=150)
        self.tree_returns.column("date", width=150)
        self.tree_returns.column("invoice", width=120)
        self.tree_returns.column("total", width=80)
        self.tree_returns.column("by", width=80)
        self.tree_returns.pack(fill="both", expand=True, padx=8, pady=2)
        self.tree_returns.bind("<Double-1>", self._show_return_details_dialog)

    def _return_search_key(self, event=None):
        try:
            q = self.return_search_var.get().strip()
            self.lb_return_suggest.delete(0, tk.END)
            self._returns_reload(search_q=q)
            if not q:
                for i in self.tree_invoice_preview.get_children():
                    self.tree_invoice_preview.delete(i)
                return
            self._inv_sug_rows = self.db.search_invoice_numbers(q, limit=5)
            for r in self._inv_sug_rows:
                self.lb_return_suggest.insert(tk.END, f"{r['invoice_number']}")
            self._preview_invoice_content()
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء البحث: {e}")

    def _return_suggest_select(self, event=None):
        try:
            sel = self.lb_return_suggest.curselection()
            if not sel:
                return
            idx = sel[0]
            if not hasattr(self, '_inv_sug_rows') or not self._inv_sug_rows:
                return
            chosen_inv = self._inv_sug_rows[idx]['invoice_number']
            self.return_search_var.set(chosen_inv)
            self.lb_return_suggest.delete(0, tk.END)
            self._returns_reload(search_q=chosen_inv)
            self._preview_invoice_content()
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء الاختيار: {e}")

    def _preview_invoice_content(self):
        try:
            for i in self.tree_invoice_preview.get_children():
                self.tree_invoice_preview.delete(i)
            invno = self.return_search_var.get().strip()
            if not invno:
                return
            # استخدام LIKE للسماح بالبحث الجزئي
            r = self.db.conn.execute(
                "SELECT id FROM invoices WHERE invoice_number LIKE ? ORDER BY invoice_date DESC LIMIT 1",
                (f"%{invno}%",)
            ).fetchone()
            if not r:
                return
            # ✅ استخدام get_invoice_remaining لعرض الكمية المتبقية بعد الاسترجاع
            _, lines = self.db.get_invoice_remaining(r["id"])
            for l in lines:
                self.tree_invoice_preview.insert("", tk.END, iid=str(l['medicine_id']),
                    values=(l["barcode"] or "N/A", l["name"], l['remaining_qty'], f"{l['unit_price']:.2f}", f"{l['line_total']:.2f}"))
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء استعراض الفاتورة: {e}")

    def _full_return(self):
        try:
            invno = self.return_search_var.get().strip()
            if not invno:
                messagebox.showwarning("تنبيه", "أدخل رقم الفاتورة أولاً")
                return
            r = self.db.conn.execute("SELECT id FROM invoices WHERE invoice_number=?", (invno,)).fetchone()
            if not r:
                messagebox.showerror("غير موجود", "الفاتورة غير موجودة")
                return
            inv_id = r["id"]
            if self.db.has_return_for_invoice(inv_id):
                messagebox.showerror("مسترجعة", "هذه الفاتورة قد أُرجعت بالكامل من قبل")
                return
            if not messagebox.askyesno("تأكيد", f"هل تريد استرجاع الفاتورة {invno} بالكامل؟"):
                return
            rid = self.db.create_return_full(inv_id, "local")
            if rid:
                messagebox.showinfo("تم", f"تم الاسترجاع الكامل بنجاح #{rid}")
                self.return_search_var.set("")
                self.lb_return_suggest.delete(0, tk.END)
                for i in self.tree_invoice_preview.get_children():
                    self.tree_invoice_preview.delete(i)
                self._returns_reload()
                self._stock_reload()
                self._check_alerts()
            else:
                messagebox.showerror("خطأ", "فشل الاسترجاع: الفاتورة فارغة أو مسترجعة")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء الاسترجاع الكامل: {e}")

    def _partial_return(self):
        try:
            invno = self.return_search_var.get().strip()
            focused_item = self.tree_invoice_preview.focus()
            if not invno:
                messagebox.showwarning("تنبيه", "أدخل رقم الفاتورة أولاً!")
                return
            if not focused_item:
                messagebox.showwarning("تنبيه", "حدد الصنف من جدول المعاينة أولاً!")
                return
            r = self.db.conn.execute("SELECT id FROM invoices WHERE invoice_number=?", (invno,)).fetchone()
            if not r:
                messagebox.showerror("غير موجود", "الفاتورة غير موجودة")
                return
            inv_id = r["id"]

            med_id = int(focused_item)
            # ✅ استخدام get_invoice_remaining للحصول على الكمية المتبقية الفعلية
            _, lines = self.db.get_invoice_remaining(inv_id)
            line = None
            for l in lines:
                if l['medicine_id'] == med_id:
                    line = l
                    break
            if not line:
                messagebox.showerror("خطأ", "الصنف المختار غير موجود في هذه الفاتورة أو تم استرجاعه بالكامل")
                return

            remaining_qty = line["remaining_qty"]
            unit_price = line["unit_price"]
            med_name = line["name"]

            if remaining_qty <= 0:
                messagebox.showwarning("تنبيه", "هذا الصنف تم استرجاعه بالكامل")
                return

            ret_qty = simpledialog.askinteger("استرجاع جزئي",
                f"الدواء: {med_name}\nالكمية المتبقية: {remaining_qty}\n\nأدخل الكمية المرتجعة:",
                initialvalue=1, minvalue=1, maxvalue=remaining_qty, parent=self)
            if ret_qty is None:
                return

            rid = self.db.create_return_partial(inv_id, [(med_id, ret_qty, unit_price)], "local")
            if rid:
                messagebox.showinfo("تم", f"تم الاسترجاع الجزئي بنجاح #{rid}")
                self._returns_reload(search_q=invno)
                self._preview_invoice_content()
                self._stock_reload()
                self._check_alerts()
            else:
                messagebox.showerror("فشل", "فشل إنشاء الاسترجاع")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء الاسترجاع الجزئي: {e}")

    def _show_return_details_dialog(self, event=None):
        try:
            r_id = safe_get_id(self.tree_returns.selection(), self.tree_returns)
            if not r_id:
                return
            lines = self.db.get_return_details(r_id)
            if not lines:
                return
            txt = f"إيصال رقم: {r_id}\n-----------------------\n"
            for l in lines:
                txt += f"- {l['name']} | الكمية: {l['qty']} | السعر: {l['unit_price']:.2f}\n"
            messagebox.showinfo("الأصناف المسترجعة", txt)
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ: {e}")

    def _returns_reload(self, search_q=None):
        try:
            for i in self.tree_returns.get_children():
                self.tree_returns.delete(i)
            for r in self.db.list_returns(search_q=search_q):
                self.tree_returns.insert("", tk.END, values=(
                    r["id"], r["return_number"], r["return_date"][:19],
                    r["invoice_number"] or "مستقلة", f"{r['total_amount']:.2f}", r["created_by"]))
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء تحديث المسترجعات: {e}")

    # ==================== التقارير ====================
    def _build_reports(self, parent):
        top = ttk.Frame(parent)
        top.pack(fill="x", padx=8, pady=6)

        ttk.Label(top, text="📅 من (YYYY-MM-DD):").pack(side="left")
        self.date_from_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(top, textvariable=self.date_from_var, width=11).pack(side="left", padx=2)
        ttk.Label(top, text="إلى:").pack(side="left")
        self.date_to_var = tk.StringVar(value=datetime.now().strftime("%Y-%m-%d"))
        ttk.Entry(top, textvariable=self.date_to_var, width=11).pack(side="left", padx=2)
        ttk.Button(top, text="🔍 فلترة", command=self._reports_search_date).pack(side="left", padx=4)
        ttk.Button(top, text="📋 عرض الكل", command=self._reports_reload_default).pack(side="left", padx=2)
        ttk.Button(top, text="📊 تصدير Excel", command=self._reports_export_excel).pack(side="right", padx=4)
        ttk.Button(top, text="💾 نسخة احتياطية", command=lambda: backup_db("manual")).pack(side="right", padx=4)

        self.tree_reports = ttk.Treeview(parent, columns=("id", "num", "date", "total", "by"), show="headings", height=8)
        for c, t in zip(("id", "num", "date", "total", "by"), ["#", "رقم الفاتورة", "التاريخ", "الإجمالي", "المستخدم"]):
            self.tree_reports.heading(c, text=t)
        self.tree_reports.column("id", width=40)
        self.tree_reports.column("num", width=150)
        self.tree_reports.column("date", width=150)
        self.tree_reports.column("total", width=100)
        self.tree_reports.column("by", width=100)
        self.tree_reports.pack(fill="both", expand=True, padx=8, pady=2)
        self.tree_reports.bind("<Double-1>", self._show_invoice_details)

        self.frame_financial_summary = ttk.LabelFrame(parent, text=" الحسابات المالية والأرباح ")
        self.frame_financial_summary.pack(fill="x", padx=8, pady=4)
        self.sales_label_var = tk.StringVar()
        ttk.Label(self.frame_financial_summary, textvariable=self.sales_label_var, font=("Segoe UI", 10, "bold"), foreground="#006600").pack(side="left", padx=15, pady=4)
        self.profit_label_var = tk.StringVar()
        ttk.Label(self.frame_financial_summary, textvariable=self.profit_label_var, font=("Segoe UI", 10, "bold"), foreground="#0033cc").pack(side="right", padx=15, pady=4)

    def _reports_reload_default(self):
        try:
            for i in self.tree_reports.get_children():
                self.tree_reports.delete(i)
            for r in self.db.list_all_invoices():
                self.tree_reports.insert("", tk.END, values=(
                    r["id"], r["invoice_number"], r["invoice_date"][:19], f"{r['total_amount']:.2f}", r["created_by"]))
            tsales, nprofit = self.db.calculate_financial_profits()
            self.sales_label_var.set(f"المبيعات: {tsales:.2f} {CURRENCY}")
            self.profit_label_var.set(f"الأرباح: {nprofit:.2f} {CURRENCY}")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء تحميل التقارير: {e}")

    def _reports_export_excel(self):
        """تصدير التقارير الحالية إلى ملف Excel"""
        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        except ImportError:
            messagebox.showerror("مكتبة ناقصة", "يرجى تثبيت مكتبة openpyxl: pip install openpyxl")
            return

        try:
            # الحصول على البيانات المعروضة حالياً
            rows_data = []
            for item_id in self.tree_reports.get_children():
                vals = self.tree_reports.item(item_id, "values")
                rows_data.append(vals)

            if not rows_data:
                messagebox.showwarning("فارغ", "لا توجد بيانات للتصدير")
                return

            # اختيار مكان الحفظ
            from tkinter import filedialog
            fname = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
                initialfile=f"تقرير_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            if not fname:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "التقارير"
            ws.sheet_view.rightToLeft = True  # اتجاه RTL

            # العناوين
            headers = ["#", "رقم الفاتورة", "التاريخ", "الإجمالي", "المستخدم"]
            header_fill = PatternFill(start_color="115e59", end_color="115e59", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, size=11)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )

            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=h)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

            # البيانات
            for row_idx, row_data in enumerate(rows_data, 2):
                for col_idx, val in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=val)
                    cell.alignment = Alignment(horizontal="right" if col_idx in [2,5] else "center")
                    cell.border = thin_border
                    if row_idx % 2 == 0:
                        cell.fill = PatternFill(start_color="f1f5f9", end_color="f1f5f9", fill_type="solid")

            # تعديل عرض الأعمدة
            ws.column_dimensions['A'].width = 8
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 20
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15

            # صف الملخص المالي
            next_row = len(rows_data) + 3
            ws.cell(row=next_row, column=1, value="الملخص المالي:").font = Font(bold=True, size=12, color="115e59")
            ws.cell(row=next_row + 1, column=1, value=self.sales_label_var.get()).font = Font(bold=True, size=11)
            ws.cell(row=next_row + 2, column=1, value=self.profit_label_var.get()).font = Font(bold=True, size=11, color="006600")

            wb.save(fname)
            messagebox.showinfo("تم", f"تم تصدير التقرير بنجاح: {fname}")
        except Exception as e:
            messagebox.showerror("خطأ", f"فشل التصدير: {e}")

    def _reports_search_date(self):
        try:
            fd = self.date_from_var.get().strip()
            td = self.date_to_var.get().strip()
            if not fd or not td:
                return
            try:
                datetime.strptime(fd, "%Y-%m-%d")
                datetime.strptime(td, "%Y-%m-%d")
            except:
                messagebox.showerror("خطأ", "صيغة التاريخ غير صحيحة")
                return
            fd_s, td_s = fd + "T00:00:00", td + "T23:59:59"
            for i in self.tree_reports.get_children():
                self.tree_reports.delete(i)
            for r in self.db.list_invoices_between(fd_s, td_s):
                self.tree_reports.insert("", tk.END, values=(
                    r["id"], r["invoice_number"], r["invoice_date"][:19], f"{r['total_amount']:.2f}", r["created_by"]))
            tsales, nprofit = self.db.calculate_financial_profits(fd_s, td_s)
            self.sales_label_var.set(f"المبيعات: {tsales:.2f} {CURRENCY}")
            self.profit_label_var.set(f"الأرباح: {nprofit:.2f} {CURRENCY}")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء الفلترة: {e}")

    def _show_invoice_details(self, event=None):
        try:
            inv_id = safe_get_id(self.tree_reports.selection(), self.tree_reports)
            if not inv_id:
                return
            inv, lines = self.db.get_invoice(inv_id)
            if not inv:
                return
            txt = f"فاتورة: {inv['invoice_number']}\nالتاريخ: {inv['invoice_date']}\nالإجمالي: {inv['total_amount']:.2f} {inv['currency']}\nالأصناف:\n"
            for l in lines:
                txt += f"- {l['name']} | الكمية: {l['qty']} | السعر: {l['unit_price']:.2f}\n"
            messagebox.showinfo("تفاصيل الفاتورة", txt)
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء عرض التفاصيل: {e}")

    # ==================== الأطباء ====================
    def _build_doctors_tab(self, parent):
        left = ttk.Frame(parent)
        left.pack(side="left", fill="y", padx=8, pady=8)

        ttk.Label(left, text="اسم الطبيب:").pack(anchor="w")
        self.doc_name_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.doc_name_var).pack(fill="x", pady=2)
        ttk.Label(left, text="التخصص:").pack(anchor="w")
        self.doc_spec_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.doc_spec_var).pack(fill="x", pady=2)
        ttk.Label(left, text="الهاتف:").pack(anchor="w")
        self.doc_phone_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.doc_phone_var).pack(fill="x", pady=2)

        ttk.Button(left, text="➕ إضافة طبيب", command=self._doctor_save).pack(fill="x", pady=4)
        ttk.Button(left, text="✏️ تعديل الطبيب", command=self._doctor_edit).pack(fill="x", pady=2)
        ttk.Button(left, text="🗑️ حذف الطبيب", command=self._doctor_delete).pack(fill="x", pady=2)

        right = ttk.Frame(parent)
        right.pack(side="right", fill="both", expand=True, padx=8, pady=8)

        self.tree_doctors = ttk.Treeview(right, columns=("id", "name", "specialty", "phone"), show="headings", height=10)
        for c, t in zip(("id", "name", "specialty", "phone"), ["#", "اسم الطبيب", "التخصص", "الهاتف"]):
            self.tree_doctors.heading(c, text=t)
        self.tree_doctors.column("id", width=40)
        self.tree_doctors.column("name", width=200)
        self.tree_doctors.column("specialty", width=150)
        self.tree_doctors.column("phone", width=120)
        self.tree_doctors.pack(fill="both", expand=True)
        self.tree_doctors.bind("<Double-1>", self._doctor_load_to_entries)
        self._selected_doc_id = None

    def _doctor_load_to_entries(self, event=None):
        sel = self.tree_doctors.selection()
        if not sel:
            return
        vals = self.tree_doctors.item(sel, "values")
        if not vals:
            return
        self._selected_doc_id = int(vals[0])
        self.doc_name_var.set(vals[1])
        self.doc_spec_var.set(vals[2])
        self.doc_phone_var.set(vals[3])

    def _doctor_save(self):
        n = self.doc_name_var.get().strip()
        s = self.doc_spec_var.get().strip()
        p = self.doc_phone_var.get().strip()
        if not n:
            messagebox.showwarning("تنبيه", "اسم الطبيب مطلوب!")
            return
        self.db.add_doctor(n, s, p)
        messagebox.showinfo("تم", "تم الحفظ بنجاح")
        self._doctors_reload()
        self.doc_name_var.set("")
        self.doc_spec_var.set("")
        self.doc_phone_var.set("")
        self._selected_doc_id = None

    def _doctor_edit(self):
        if not self._selected_doc_id:
            messagebox.showwarning("تنبيه", "اختر طبيباً أولاً!")
            return
        self.db.update_doctor(self._selected_doc_id, self.doc_name_var.get().strip(),
                              self.doc_spec_var.get().strip(), self.doc_phone_var.get().strip())
        messagebox.showinfo("تم", "تم التعديل بنجاح")
        self._doctors_reload()

    def _doctor_delete(self):
        if not self._selected_doc_id:
            messagebox.showwarning("تنبيه", "اختر طبيباً أولاً!")
            return
        if messagebox.askyesno("تأكيد", "هل تريد حذف هذا الطبيب؟"):
            self.db.delete_doctor(self._selected_doc_id)
            self._doctors_reload()
            self.doc_name_var.set("")
            self.doc_spec_var.set("")
            self.doc_phone_var.set("")
            self._selected_doc_id = None

    def _doctors_reload(self):
        for i in self.tree_doctors.get_children():
            self.tree_doctors.delete(i)
        for r in self.db.list_doctors():
            self.tree_doctors.insert("", tk.END, values=(r["id"], r["name"], r["specialty"], r["phone"]))
        try:
            self.pos_tab.refresh_clinical_combos()
        except:
            pass

    # ==================== المرضى ====================
    def _build_patients_tab(self, parent):
        left = ttk.Frame(parent)
        left.pack(side="left", fill="y", padx=8, pady=8)

        ttk.Label(left, text="اسم المريض:").pack(anchor="w")
        self.pat_name_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.pat_name_var).pack(fill="x", pady=2)
        ttk.Label(left, text="العمر:").pack(anchor="w")
        self.pat_age_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.pat_age_var).pack(fill="x", pady=2)
        ttk.Label(left, text="الجنس:").pack(anchor="w")
        self.pat_gender_var = tk.StringVar(value="ذكر")
        ttk.Combobox(left, textvariable=self.pat_gender_var, values=["ذكر", "أنثى"], state="readonly").pack(fill="x", pady=2)
        ttk.Label(left, text="رقم الهاتف:").pack(anchor="w")
        self.pat_phone_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.pat_phone_var).pack(fill="x", pady=2)

        ttk.Button(left, text="➕ تسجيل مريض جديد", command=self._patient_save).pack(fill="x", pady=4)
        ttk.Button(left, text="✏️ تعديل المريض", command=self._patient_edit).pack(fill="x", pady=2)
        ttk.Button(left, text="🗑️ حذف المريض", command=self._patient_delete).pack(fill="x", pady=2)
        ttk.Button(left, text="📋 عرض السجل الطبي", command=self._show_patient_history).pack(fill="x", pady=4)

        right = ttk.Frame(parent)
        right.pack(side="right", fill="both", expand=True, padx=8, pady=8)

        self.tree_patients = ttk.Treeview(right, columns=("id", "name", "age", "gender", "phone", "visits"), show="headings", height=10)
        for c, t in zip(("id", "name", "age", "gender", "phone", "visits"), ["#", "اسم المريض", "العمر", "الجنس", "الهاتف", "الزيارات"]):
            self.tree_patients.heading(c, text=t)
        self.tree_patients.column("id", width=40)
        self.tree_patients.column("name", width=180)
        self.tree_patients.column("age", width=50, anchor="center")
        self.tree_patients.column("gender", width=60, anchor="center")
        self.tree_patients.column("phone", width=100)
        self.tree_patients.column("visits", width=60, anchor="center")
        self.tree_patients.pack(fill="both", expand=True)
        self.tree_patients.bind("<Double-1>", self._show_patient_history)
        self.tree_patients.bind("<<TreeviewSelect>>", self._patient_load_to_entries)
        self._selected_pat_id = None

    def _patient_load_to_entries(self, event=None):
        sel = self.tree_patients.selection()
        if not sel:
            return
        vals = self.tree_patients.item(sel, "values")
        if not vals or len(vals) < 6:
            return
        self._selected_pat_id = int(vals[0])
        self.pat_name_var.set(vals[1])
        self.pat_age_var.set(str(vals[2]))
        self.pat_gender_var.set(vals[3])
        self.pat_phone_var.set(vals[4])

    def _patient_save(self):
        n = self.pat_name_var.get().strip()
        a = self.pat_age_var.get().strip()
        g = self.pat_gender_var.get()
        p = self.pat_phone_var.get().strip()
        if not n:
            messagebox.showwarning("تنبيه", "اسم المريض مطلوب!")
            return
        try:
            age = int(a) if a else 0
        except:
            messagebox.showerror("خطأ", "العمر يجب أن يكون رقماً")
            return
        self.db.add_patient(n, age, g, p)
        messagebox.showinfo("تم", "تم التسجيل بنجاح")
        self._patients_reload()
        self.pat_name_var.set("")
        self.pat_age_var.set("")
        self.pat_phone_var.set("")
        self._selected_pat_id = None

    def _patient_edit(self):
        if not self._selected_pat_id:
            messagebox.showwarning("تنبيه", "اختر مريضاً أولاً!")
            return
        try:
            age = int(self.pat_age_var.get())
        except:
            age = 0
        self.db.update_patient(self._selected_pat_id, self.pat_name_var.get().strip(),
                               age, self.pat_gender_var.get(), self.pat_phone_var.get().strip())
        messagebox.showinfo("تم", "تم تعديل المريض")
        self._patients_reload()

    def _patient_delete(self):
        if not self._selected_pat_id:
            messagebox.showwarning("تنبيه", "اختر مريضاً أولاً!")
            return
        if messagebox.askyesno("تأكيد", "هل تريد حذف المريض وسجلاته؟"):
            self.db.delete_patient(self._selected_pat_id)
            self._patients_reload()
            self.pat_name_var.set("")
            self.pat_age_var.set("")
            self.pat_phone_var.set("")
            self._selected_pat_id = None

    def _patients_reload(self):
        for i in self.tree_patients.get_children():
            self.tree_patients.delete(i)
        for r in self.db.list_patients():
            self.tree_patients.insert("", tk.END, values=(
                r["id"], r["name"], r["age"], r["gender"], r["phone"], r["visits_count"]))
        try:
            self.pos_tab.refresh_clinical_combos()
        except:
            pass

    def _show_patient_history(self, event=None):
        try:
            sel = self.tree_patients.selection()
            p_id = safe_get_id(sel, self.tree_patients)
            if not p_id:
                messagebox.showwarning("تنبيه", "اختر مريضاً أولاً")
                return
            vals = self.tree_patients.item(sel, "values")
            p_name = vals[1] if vals and len(vals) > 1 else f"مريض #{p_id}"
            records, medicines = self.db.get_patient_history(p_id)

            history_win = tk.Toplevel(self)
            history_win.title(f"الملف الطبي: {p_name}")
            history_win.geometry("650x550")

            txt_area = tk.Text(history_win, wrap="word", font=("Segoe UI", 10))
            scrollbar = ttk.Scrollbar(history_win, orient="vertical", command=txt_area.yview)
            txt_area.configure(yscrollcommand=scrollbar.set)
            txt_area.pack(side="left", fill="both", expand=True, padx=10, pady=10)
            scrollbar.pack(side="right", fill="y")

            report_text = f"📋 الملف الطبي: {p_name} (رقم الملف: {p_id})\n{'='*50}\n\n🔬 أولاً: التشخيصات والزيارات الطبية:\n"
            for r in records:
                report_text += f"\n📅 التاريخ: {r['visit_date']} | 👨‍⚕️ الطبيب: {r['doctor_name'] or 'غير محدد'}\n"
                report_text += f"🔍 التشخيص: {r['diagnosis'] or 'لا يوجد'}\n"
                report_text += f"📝 ملاحظات: {r['notes'] or 'لا يوجد'}\n"
                report_text += "-" * 40 + "\n"

            report_text += "\n\n💊 ثانياً: الأدوية ووصف استخدام العلاج:\n"
            for m in medicines:
                report_text += f"\n📅 التاريخ: {m['invoice_date'][:16]}\n"
                report_text += f"📦 الدواء: {m['med_name']} | الكمية: {m['qty']}\n"
                report_text += f"💡 الجرعة: {m['dosage_instruction'] or 'حسب الإرشادات'}\n"
                report_text += "-" * 40 + "\n"

            txt_area.insert("1.0", report_text)
            txt_area.config(state="disabled")
        except Exception as e:
            messagebox.showerror("خطأ", f"حدث خطأ أثناء عرض السجل: {e}")

    # ==================== المستخدمين ====================
    def _build_users_tab(self, parent):
        left = ttk.Frame(parent)
        left.pack(side="left", fill="y", padx=8, pady=8)

        ttk.Label(left, text="اسم المستخدم:").pack(anchor="w")
        self.u_name_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.u_name_var).pack(fill="x", pady=2)
        ttk.Label(left, text="كلمة المرور الجديدة:").pack(anchor="w")
        self.u_pass_var = tk.StringVar()
        ttk.Entry(left, textvariable=self.u_pass_var, show="*").pack(fill="x", pady=2)
        ttk.Label(left, text="الصلاحية:").pack(anchor="w")
        self.u_role_var = tk.StringVar(value="cashier")
        ttk.Combobox(left, textvariable=self.u_role_var, values=["admin", "cashier"], state="readonly").pack(fill="x", pady=2)

        ttk.Button(left, text="💾 حفظ / تحديث", command=self._user_save).pack(fill="x", pady=10)
        ttk.Button(left, text="🗑️ حذف مستخدم", command=self._user_delete).pack(fill="x", pady=2)

        right = ttk.Frame(parent)
        right.pack(side="right", fill="both", expand=True, padx=8, pady=8)

        self.tree_users = ttk.Treeview(right, columns=("id", "username", "role"), show="headings")
        for c, t in zip(("id", "username", "role"), ["#", "اسم المستخدم", "الصلاحية"]):
            self.tree_users.heading(c, text=t)
        self.tree_users.column("id", width=40)
        self.tree_users.column("username", width=200)
        self.tree_users.column("role", width=100)
        self.tree_users.pack(fill="both", expand=True)
        self.tree_users.bind("<Double-1>", self._user_load_to_entries)
        self._selected_user_id = None

    def _user_load_to_entries(self, event=None):
        sel = self.tree_users.selection()
        if not sel:
            return
        vals = self.tree_users.item(sel, "values")
        if not vals or len(vals) < 3:
            return
        self._selected_user_id = int(vals[0])
        self.u_name_var.set(vals[1])
        self.u_role_var.set(vals[2])
        self.u_pass_var.set("")

    def _user_save(self):
        u = self.u_name_var.get().strip()
        p = self.u_pass_var.get().strip()
        r = self.u_role_var.get()
        if not u:
            messagebox.showwarning("تنبيه", "اسم المستخدم مطلوب!")
            return
        if not p and not self._selected_user_id:
            messagebox.showwarning("تنبيه", "كلمة المرور مطلوبة للمستخدم الجديد!")
            return
        # إذا كان تعديل ولم يُدخل كلمة مرور جديدة، احتفظ بالقديمة
        if not p and self._selected_user_id:
            cur = self.db.conn.cursor()
            cur.execute("SELECT password FROM users WHERE id=?", (self._selected_user_id,))
            row = cur.fetchone()
            if row:
                p = row["password"]
        self.db.add_or_update_user(u, p, r)
        messagebox.showinfo("تم", "تم الحفظ بنجاح")
        self._users_reload()
        self.u_name_var.set("")
        self.u_pass_var.set("")
        self.u_role_var.set("cashier")
        self._selected_user_id = None

    def _user_delete(self):
        if not self._selected_user_id:
            messagebox.showwarning("تنبيه", "اختر مستخدماً أولاً!")
            return
        cur = self.db.conn.cursor()
        cur.execute("SELECT username FROM users WHERE id=?", (self._selected_user_id,))
        row = cur.fetchone()
        if row and row["username"] == getattr(self, "current_user_name", ""):
            messagebox.showerror("خطأ", "لا يمكنك حذف حسابك الحالي!")
            return
        if messagebox.askyesno("تأكيد", "هل تريد حذف هذا المستخدم؟"):
            self.db.delete_user(self._selected_user_id)
            self._users_reload()
            self.u_name_var.set("")
            self.u_pass_var.set("")
            self.u_role_var.set("cashier")
            self._selected_user_id = None

    def _users_reload(self):
        for i in self.tree_users.get_children():
            self.tree_users.delete(i)
        for r in self.db.list_all_users():
            self.tree_users.insert("", tk.END, values=(r["id"], r["username"], r["role"]))

    # ==================== التنبيهات ====================
    def _build_alerts(self, parent):
        top = ttk.Frame(parent)
        top.pack(fill="x", padx=8, pady=6)
        ttk.Button(top, text="🔄 تحديث التنبيهات", command=self._check_alerts).pack(side="left", padx=4)

        # منخفض المخزون
        low_frame = ttk.LabelFrame(parent, text="📉 منخفض المخزون")
        low_frame.pack(fill="both", expand=True, padx=8, pady=4)
        cols_low = ("barcode", "name", "qty", "min_limit")
        self.tree_low_stock = ttk.Treeview(low_frame, columns=cols_low, show="headings", height=5)
        for c, t in zip(cols_low, ["باركود", "الاسم", "الكمية", "الحد الأدنى"]):
            self.tree_low_stock.heading(c, text=t)
        self.tree_low_stock.pack(fill="both", expand=True, padx=4, pady=4)

        # قارب على الانتهاء
        exp_frame = ttk.LabelFrame(parent, text="⏳ قارب على الانتهاء (30 يوم)")
        exp_frame.pack(fill="both", expand=True, padx=8, pady=4)
        cols_exp = ("barcode", "name", "expiry", "days_left")
        self.tree_near_expiry = ttk.Treeview(exp_frame, columns=cols_exp, show="headings", height=5)
        for c, t in zip(cols_exp, ["باركود", "الاسم", "تاريخ الانتهاء", "الأيام المتبقية"]):
            self.tree_near_expiry.heading(c, text=t)
        self.tree_near_expiry.pack(fill="both", expand=True, padx=4, pady=4)

        # منتهي
        expired_frame = ttk.LabelFrame(parent, text="🔴 منتهي الصلاحية")
        expired_frame.pack(fill="both", expand=True, padx=8, pady=4)
        cols_ex = ("barcode", "name", "expiry")
        self.tree_expired = ttk.Treeview(expired_frame, columns=cols_ex, show="headings", height=4)
        for c, t in zip(cols_ex, ["باركود", "الاسم", "تاريخ الانتهاء"]):
            self.tree_expired.heading(c, text=t)
        self.tree_expired.pack(fill="both", expand=True, padx=4, pady=4)

    def _check_alerts(self):
        try:
            # منخفض المخزون
            for i in self.tree_low_stock.get_children():
                self.tree_low_stock.delete(i)
            low_meds = self.db.get_low_stock_medicines()
            for r in low_meds:
                self.tree_low_stock.insert("", tk.END, values=(r["barcode"], r["name"], r["quantity"], r["min_limit"]))

            # قارب على الانتهاء
            for i in self.tree_near_expiry.get_children():
                self.tree_near_expiry.delete(i)
            near_meds = self.db.get_near_expiry(days=30)
            for r in near_meds:
                try:
                    exp = datetime.strptime(r["expiry_date"], "%Y-%m-%d").date()
                    days = (exp - datetime.now().date()).days
                except:
                    days = "?"
                self.tree_near_expiry.insert("", tk.END, values=(r["barcode"], r["name"], r["expiry_date"], days))

            # منتهي
            for i in self.tree_expired.get_children():
                self.tree_expired.delete(i)
            expired_meds = self.db.get_expired_medicines()
            for r in expired_meds:
                self.tree_expired.insert("", tk.END, values=(r["barcode"], r["name"], r["expiry_date"]))

            # تحديث شريط الحالة
            low_count = len(low_meds)
            exp_count = len(near_meds) + len(expired_meds)
            if low_count > 0 or exp_count > 0:
                self.status_var.set(f"⚠️ تنبيهات: {low_count} منخفض المخزون | {exp_count} صلاحية | المستخدم: {getattr(self, 'current_user_name', 'unknown')}")
            else:
                self.status_var.set(f"✅ لا توجد تنبيهات | المستخدم: {getattr(self, 'current_user_name', 'unknown')}")
        except Exception as e:
            print(f"Alert check error: {e}")

    def on_close(self):
        try:
            self.db.close()
        except:
            pass
        self.destroy()

if __name__ == "__main__":
    app = PharmacyApp()
    app.protocol("WM_DELETE_WINDOW", app.on_close)
    app.mainloop()