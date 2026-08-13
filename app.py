from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_from_directory
from functools import wraps
from datetime import datetime, timedelta
import os

from config import CURRENCY, CURRENCY_SYMBOL, PDF_DIR, SYNC_CONFIG
from pharmacy_db import update_database, backup_db, list_backups, restore_db, delete_backup, BACKUP_DIR
from pharmacy_backend import DB
from sync_manager import SyncManager

app = Flask(__name__)
app.secret_key = "pharmacy-hybrid-secret-key-change-in-production"

update_database()
db = DB()
sync_manager = SyncManager(db.conn)
if SYNC_CONFIG["sync_on_startup"]:
    sync_manager.start_auto_sync()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated

def admin_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        if session.get('role') != 'admin':
            flash('غير مصرح لك بالوصول', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated

@app.before_request
def check_license():
    allowed = ['login', 'license', 'static']
    if request.endpoint in allowed:
        return
    if 'license_valid' not in session:
        is_valid, msg = db.check_license()
        if not is_valid:
            session['license_msg'] = msg
            return redirect(url_for('license'))
        session['license_valid'] = True

@app.route('/license', methods=['GET', 'POST'])
def license():
    if request.method == 'POST':
        code = request.form.get('code', '').strip()
        ok, msg = db.activate_license(code)
        if ok:
            session['license_valid'] = True
            flash(msg, 'success')
            return redirect(url_for('login'))
        flash(msg, 'danger')
    return render_template('license.html', status_msg=session.get('license_msg', 'البرنامج يتطلب تفعيلاً'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = request.form.get('username', '').strip()
        p = request.form.get('password', '').strip()
        user = db.check_login(u, p)
        if user:
            session['user_id'] = user['id']
            session['username'] = user['username']
            session['role'] = user['role']
            flash(f'مرحباً بك {user["username"]}', 'success')
            return redirect(url_for('dashboard'))
        flash('بيانات الدخول غير صحيحة', 'danger')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/')
@app.route('/dashboard')
@login_required
def dashboard():
    medicines = db.list_medicines()
    invoices = db.list_all_invoices()[:10]
    sales, profit = db.calculate_financial_profits()
    low = db.get_low_stock_medicines()
    expired = db.get_expired_medicines()
    stats = {
        'medicines': len(medicines),
        'invoices': len(db.list_all_invoices()),
        'sales': sales,
        'profit': profit,
        'low_stock': len(low),
        'expired': len(expired)
    }
    sync_status = sync_manager.get_sync_status()
    return render_template('dashboard.html', stats=stats, recent_invoices=invoices,
                           currency=CURRENCY, sync_status=sync_status,
                           active_page='dashboard', now=datetime.now().strftime('%Y-%m-%d'))

@app.route('/pos')
@login_required
def pos():
    patients = db.list_patients()
    doctors = db.list_doctors()
    return render_template('pos.html', patients=patients, doctors=doctors,
                           currency_symbol=CURRENCY_SYMBOL.get(CURRENCY, ''), active_page='pos')

@app.route('/api/search_medicines')
@login_required
def api_search_medicines():
    q = request.args.get('q', '')
    return jsonify([dict(r) for r in db.search_medicines(q, limit=50)])

@app.route('/api/medicine_by_barcode')
@login_required
def api_medicine_by_barcode():
    code = request.args.get('barcode', '')
    med = db.get_medicine_by_barcode(code)
    if med:
        return jsonify({'found': True, 'medicine': dict(med)})
    return jsonify({'found': False})

@app.route('/api/checkout', methods=['POST'])
@login_required
def api_checkout():
    data = request.get_json()
    items = data.get('items', [])
    patient_id = data.get('patient_id')
    doctor_id = data.get('doctor_id')
    discount = float(data.get('discount', 0) or 0)
    if not items:
        return jsonify({'success': False, 'message': 'الفاتورة فارغة'})
    for it in items:
        med = db.get_medicine_by_id(it['medicine_id'])
        if not med or med['quantity'] < it['qty']:
            return jsonify({'success': False, 'message': f"الكمية غير كافية: {it['name']}"})
    invnum = f"INV-{datetime.now().strftime('%Y%m%d')}-{int(datetime.now().timestamp())%10000}"
    if discount > 0:
        total = sum(i['line_total'] for i in items)
        ratio = max(0, (total - discount) / total) if total > 0 else 0
        for i in items:
            i['line_total'] = i['line_total'] * ratio
            i['unit_price'] = i['unit_price'] * ratio
    inv_id = db.save_invoice_clinical(invnum, items, CURRENCY, session['username'], patient_id, doctor_id)
    return jsonify({'success': True, 'invoice_number': invnum, 'invoice_id': inv_id})

@app.route('/api/print_preview', methods=['POST'])
@login_required
def api_print_preview():
    data = request.get_json()
    items = data.get('items', [])
    discount = float(data.get('discount', 0) or 0)
    subtotal = sum(i['line_total'] for i in items)
    total = max(0, subtotal - discount)
    sym = CURRENCY_SYMBOL.get(CURRENCY, '')
    rows = ""
    for it in items:
        rows += f"<tr><td>{it['name']}</td><td style='text-align:center'>{it['qty']}</td><td style='text-align:right'>{it['unit_price']:.2f}{sym}</td><td style='text-align:right'>{it['line_total']:.2f}{sym}</td><td>{it.get('dosage_instruction','')}</td></tr>"
    html = f"""<!DOCTYPE html>
<html dir="rtl" lang="ar">
<head><meta charset="utf-8"><style>
body{{font-family:'Segoe UI',Arial,sans-serif;direction:rtl;padding:20px;background:#f8fafc}}
.header{{text-align:center;margin-bottom:20px}} .header h2{{color:#115e59;margin:0}}
table{{width:100%;border-collapse:collapse;background:#fff;box-shadow:0 1px 3px rgba(0,0,0,0.1)}}
th,td{{border:1px solid #e2e8f0;padding:10px;text-align:right}}
th{{background:#115e59;color:#fff;font-weight:bold}}
tr:nth-child(even){{background:#f1f5f9}}
.total-row{{background:#ecfdf5;font-weight:bold;font-size:16px}}
.total-row td{{color:#065f46}}
</style></head>
<body>
<div class="header"><h2>🧾 معاينة الفاتورة</h2><p>تاريخ: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p></div>
<table>
<thead><tr><th>الصنف</th><th>الكمية</th><th>السعر</th><th>الإجمالي</th><th>الإرشادات</th></tr></thead>
<tbody>{rows}</tbody>
<tfoot>
<tr class="total-row"><td colspan="3" style="text-align:left">المجموع:</td><td style="text-align:right">{subtotal:.2f}{sym}</td><td></td></tr>
<tr class="total-row"><td colspan="3" style="text-align:left">الخصم:</td><td style="text-align:right">{discount:.2f}{sym}</td><td></td></tr>
<tr class="total-row"><td colspan="3" style="text-align:left;font-size:18px">الصافي:</td><td style="text-align:right;font-size:18px">{total:.2f}{sym}</td><td></td></tr>
</tfoot>
</table>
</body></html>"""
    fname = os.path.join(PDF_DIR, f"invoice_preview_{datetime.now().strftime('%Y%m%d_%H%M%S')}.html")
    with open(fname, "w", encoding="utf-8") as f:
        f.write(html)
    return jsonify({'url': f'/invoices_pdf/{os.path.basename(fname)}'})

@app.route('/invoices_pdf/<path:filename>')
@login_required
def serve_pdf(filename):
    return send_from_directory(PDF_DIR, filename)

@app.route('/stock', methods=['GET', 'POST'])
@login_required
@admin_required
def stock():
    today = datetime.now().strftime('%Y-%m-%d')
    near = (datetime.now() + timedelta(days=30)).strftime('%Y-%m-%d')
    if request.method == 'POST':
        edit_id = request.form.get('edit_id')
        barcode = request.form.get('barcode', '').strip()
        name = request.form.get('name', '').strip()
        sci = request.form.get('scientific_name', '').strip()
        qty = int(request.form.get('quantity', 0))
        minl = int(request.form.get('min_limit', 0))
        exp = request.form.get('expiry_date') or None
        pur = float(request.form.get('purchase_price', 0))
        sell = float(request.form.get('selling_price', 0))
        if not barcode or not name:
            flash('الباركود والاسم مطلوبان', 'danger')
        else:
            db.add_or_update_medicine(barcode, name, sci, qty, minl, exp, pur, sell)
            flash('تم الحفظ بنجاح', 'success')
            return redirect(url_for('stock'))
    search_q = request.args.get('q', '')
    medicines = db.list_medicines()
    if search_q:
        sq = search_q.lower()
        medicines = [m for m in medicines if sq in (m['name'] or '').lower()
                     or sq in (m['barcode'] or '').lower()
                     or sq in (m['scientific_name'] or '').lower()]
    edit_id = request.args.get('edit_id')
    form = {}
    if edit_id:
        med = db.get_medicine_by_id(int(edit_id))
        if med:
            form = dict(med)
    return render_template('stock.html', medicines=medicines, form=form, edit_id=edit_id,
                           search_q=search_q, today=today, near_expiry_date=near,
                           currency_symbol=CURRENCY_SYMBOL.get(CURRENCY, ''), active_page='stock')

@app.route('/stock/edit/<int:id>')
@login_required
@admin_required
def stock_edit(id):
    return redirect(url_for('stock', edit_id=id))

@app.route('/stock/delete/<int:id>')
@login_required
@admin_required
def stock_delete(id):
    ok, msg = db.delete_medicine(id)
    flash(msg, 'success' if ok else 'danger')
    return redirect(url_for('stock'))

@app.route('/stock/backup', methods=['POST'])
@login_required
@admin_required
def stock_backup():
    path = backup_db("manual_stock")
    if path:
        flash(f'تم إنشاء نسخة احتياطية: {os.path.basename(path)}', 'success')
    else:
        flash('فشل إنشاء النسخة الاحتياطية', 'danger')
    return redirect(url_for('stock'))

@app.route('/returns')
@login_required
@admin_required
def returns():
    returns_list = db.list_returns()
    return render_template('returns.html', returns=returns_list,
                           currency_symbol=CURRENCY_SYMBOL.get(CURRENCY, ''), active_page='returns')

@app.route('/api/invoice_search')
@login_required
@admin_required
def api_invoice_search():
    q = request.args.get('q', '').strip()
    if not q:
        return jsonify({'found': False, 'message': 'أدخل رقم الفاتورة'})
    inv = db.conn.execute("SELECT id, invoice_number FROM invoices WHERE invoice_number=?", (q,)).fetchone()
    if not inv:
        rows = db.search_invoice_numbers(q, limit=5)
        if not rows:
            return jsonify({'found': False, 'message': 'الفاتورة غير موجودة'})
        inv = rows[0]
    inv_id = inv['id']
    inv_data, lines = db.get_invoice_remaining(inv_id)
    if not inv_data:
        return jsonify({'found': False, 'message': 'الفاتورة غير موجودة'})
    return jsonify({'found': True, 'invoice_id': inv_id, 'invoice_number': inv_data['invoice_number'],
                    'lines': [dict(l) for l in lines]})

@app.route('/api/return_full', methods=['POST'])
@login_required
@admin_required
def api_return_full():
    data = request.get_json()
    inv_id = data.get('invoice_id')
    if not inv_id:
        return jsonify({'success': False, 'message': 'رقم الفاتورة مطلوب'})
    if db.has_return_for_invoice(inv_id):
        return jsonify({'success': False, 'message': 'هذه الفاتورة مسترجعة بالكامل مسبقاً'})
    _, lines = db.get_invoice_remaining(inv_id)
    if not lines:
        return jsonify({'success': False, 'message': 'لا توجد أصناف متبقية للاسترجاع'})
    rid = db.create_return_full(inv_id, session['username'])
    if rid:
        return jsonify({'success': True, 'message': f'تم الاسترجاع الكامل بنجاح #{rid}'})
    return jsonify({'success': False, 'message': 'فشل الاسترجاع'})

@app.route('/api/return_partial', methods=['POST'])
@login_required
@admin_required
def api_return_partial():
    data = request.get_json()
    inv_id = data.get('invoice_id')
    items = data.get('items', [])
    if not inv_id or not items:
        return jsonify({'success': False, 'message': 'بيانات ناقصة'})
    _, remaining = db.get_invoice_remaining(inv_id)
    rem_map = {l['medicine_id']: l['remaining_qty'] for l in remaining}
    for it in items:
        if rem_map.get(it['medicine_id'], 0) < it['qty']:
            return jsonify({'success': False, 'message': f'الكمية غير كافية (متاح: {rem_map.get(it["medicine_id"], 0)})'})
    parsed = [(it['medicine_id'], it['qty'], it['unit_price']) for it in items]
    rid = db.create_return_partial(inv_id, parsed, session['username'])
    if rid:
        return jsonify({'success': True, 'message': f'تم الاسترجاع الجزئي بنجاح #{rid}'})
    return jsonify({'success': False, 'message': 'فشل الاسترجاع'})

@app.route('/api/return_details')
@login_required
@admin_required
def api_return_details():
    rid = request.args.get('id', type=int)
    if not rid:
        return jsonify([])
    lines = db.get_return_details(rid)
    return jsonify([dict(l) for l in lines])

@app.route('/reports')
@login_required
@admin_required
def reports():
    date_from = request.args.get('date_from', datetime.now().strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.now().strftime('%Y-%m-%d'))
    fd_s = date_from + "T00:00:00"
    td_s = date_to + "T23:59:59"
    invoices = db.list_invoices_between(fd_s, td_s)
    sales, profit = db.calculate_financial_profits(fd_s, td_s)
    backups = list_backups()
    return render_template('reports.html', invoices=invoices, sales=sales, profit=profit,
                           date_from=date_from, date_to=date_to, currency=CURRENCY,
                           backups=backups, active_page='reports')

@app.route('/api/invoice_details')
@login_required
@admin_required
def api_invoice_details():
    inv_id = request.args.get('id', type=int)
    inv, lines = db.get_invoice(inv_id)
    return jsonify({'invoice': dict(inv), 'lines': [dict(l) for l in lines]})

@app.route('/reports/export')
@login_required
@admin_required
def reports_export():
    try:
        import openpyxl
    except ImportError:
        flash('يرجى تثبيت openpyxl', 'danger')
        return redirect(url_for('reports'))
    date_from = request.args.get('date_from', datetime.now().strftime('%Y-%m-%d'))
    date_to = request.args.get('date_to', datetime.now().strftime('%Y-%m-%d'))
    fd_s = date_from + "T00:00:00"
    td_s = date_to + "T23:59:59"
    invoices = db.list_invoices_between(fd_s, td_s)
    sales, profit = db.calculate_financial_profits(fd_s, td_s)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "التقارير"
    ws.sheet_view.rightToLeft = True
    headers = ["#", "رقم الفاتورة", "التاريخ", "الإجمالي", "المستخدم"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        cell.fill = openpyxl.styles.PatternFill(start_color="115e59", fill_type="solid")
    for row_idx, inv in enumerate(invoices, 2):
        ws.cell(row=row_idx, column=1, value=inv['id'])
        ws.cell(row=row_idx, column=2, value=inv['invoice_number'])
        ws.cell(row=row_idx, column=3, value=inv['invoice_date'][:19])
        ws.cell(row=row_idx, column=4, value=round(inv['total_amount'], 2))
        ws.cell(row=row_idx, column=5, value=inv['created_by'])
    next_row = len(invoices) + 3
    ws.cell(row=next_row, column=1, value=f"المبيعات: {sales:.2f}")
    ws.cell(row=next_row+1, column=1, value=f"الأرباح: {profit:.2f}")
    fname = f"تقرير_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(PDF_DIR, fname)
    wb.save(path)
    return send_from_directory(PDF_DIR, fname, as_attachment=True)

@app.route('/backup')
@login_required
@admin_required
def backup():
    path = backup_db("manual")
    if path:
        flash(f'تم إنشاء نسخة احتياطية: {os.path.basename(path)}', 'success')
    else:
        flash('فشل إنشاء النسخة الاحتياطية', 'danger')
    return redirect(url_for('reports'))

@app.route('/backup/download/<path:filename>')
@login_required
@admin_required
def backup_download(filename):
    return send_from_directory(BACKUP_DIR, filename, as_attachment=True)

@app.route('/backup/restore', methods=['POST'])
@login_required
@admin_required
def backup_restore():
    filename = request.form.get('filename')
    if not filename:
        flash('اختر نسخة للاستعادة', 'danger')
        return redirect(url_for('reports'))
    backup_path = os.path.join(BACKUP_DIR, filename)
    ok, msg = restore_db(backup_path)
    flash(msg, 'success' if ok else 'danger')
    if ok:
        # Restart DB connection
        global db, sync_manager
        db.close()
        db = DB()
        sync_manager = SyncManager(db.conn)
        sync_manager.start_auto_sync()
    return redirect(url_for('reports'))

@app.route('/backup/delete', methods=['POST'])
@login_required
@admin_required
def backup_delete():
    filename = request.form.get('filename')
    if not filename:
        flash('اختر نسخة للحذف', 'danger')
        return redirect(url_for('reports'))
    backup_path = os.path.join(BACKUP_DIR, filename)
    ok, msg = delete_backup(backup_path)
    flash(msg, 'success' if ok else 'danger')
    return redirect(url_for('reports'))

@app.route('/doctors', methods=['GET', 'POST'])
@login_required
@admin_required
def doctors():
    if request.method == 'POST':
        edit_id = request.form.get('edit_id')
        name = request.form.get('name', '').strip()
        spec = request.form.get('specialty', '').strip()
        phone = request.form.get('phone', '').strip()
        if not name:
            flash('اسم الطبيب مطلوب', 'danger')
        else:
            if edit_id:
                db.update_doctor(int(edit_id), name, spec, phone)
                flash('تم التعديل', 'success')
            else:
                db.add_doctor(name, spec, phone)
                flash('تم الحفظ', 'success')
            return redirect(url_for('doctors'))
    edit_id = request.args.get('edit_id')
    form = {}
    if edit_id:
        d = db.conn.execute("SELECT * FROM doctors WHERE id=?", (int(edit_id),)).fetchone()
        if d: form = dict(d)
    return render_template('doctors.html', doctors=db.list_doctors(), form=form,
                           edit_id=edit_id, active_page='doctors')

@app.route('/doctors/edit/<int:id>')
@login_required
@admin_required
def doctors_edit(id):
    return redirect(url_for('doctors', edit_id=id))

@app.route('/doctors/delete/<int:id>')
@login_required
@admin_required
def doctors_delete(id):
    db.delete_doctor(id)
    flash('تم الحذف', 'success')
    return redirect(url_for('doctors'))

@app.route('/patients', methods=['GET', 'POST'])
@login_required
@admin_required
def patients():
    if request.method == 'POST':
        edit_id = request.form.get('edit_id')
        name = request.form.get('name', '').strip()
        age = request.form.get('age', '0')
        gender = request.form.get('gender', 'ذكر')
        phone = request.form.get('phone', '').strip()
        if not name:
            flash('اسم المريض مطلوب', 'danger')
        else:
            try: age = int(age) if age else 0
            except: age = 0
            if edit_id:
                db.update_patient(int(edit_id), name, age, gender, phone)
                flash('تم التعديل', 'success')
            else:
                db.add_patient(name, age, gender, phone)
                flash('تم التسجيل', 'success')
            return redirect(url_for('patients'))
    edit_id = request.args.get('edit_id')
    form = {}
    if edit_id:
        p = db.conn.execute("SELECT * FROM patients WHERE id=?", (int(edit_id),)).fetchone()
        if p: form = dict(p)
    return render_template('patients.html', patients=db.list_patients(), form=form,
                           edit_id=edit_id, active_page='patients')

@app.route('/patients/edit/<int:id>')
@login_required
@admin_required
def patients_edit(id):
    return redirect(url_for('patients', edit_id=id))

@app.route('/patients/delete/<int:id>')
@login_required
@admin_required
def patients_delete(id):
    db.delete_patient(id)
    flash('تم الحذف', 'success')
    return redirect(url_for('patients'))

@app.route('/api/patient_history')
@login_required
@admin_required
def api_patient_history():
    pid = request.args.get('id', type=int)
    p = db.conn.execute("SELECT * FROM patients WHERE id=?", (pid,)).fetchone()
    records, medicines = db.get_patient_history(pid)
    return jsonify({'name': p['name'] if p else '',
                    'records': [dict(r) for r in records],
                    'medicines': [dict(m) for m in medicines]})

@app.route('/users', methods=['GET', 'POST'])
@login_required
@admin_required
def users():
    if request.method == 'POST':
        edit_id = request.form.get('edit_id')
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '')
        role = request.form.get('role', 'cashier')
        if not username:
            flash('اسم المستخدم مطلوب', 'danger')
        else:
            if not password and not edit_id:
                flash('كلمة المرور مطلوبة للمستخدم الجديد', 'danger')
            else:
                if not password and edit_id:
                    cur = db.conn.cursor()
                    cur.execute("SELECT password FROM users WHERE id=?", (int(edit_id),))
                    row = cur.fetchone()
                    password = row['password'] if row else ''
                db.add_or_update_user(username, password, role)
                flash('تم الحفظ', 'success')
                return redirect(url_for('users'))
    edit_id = request.args.get('edit_id')
    form = {}
    if edit_id:
        u = db.conn.execute("SELECT * FROM users WHERE id=?", (int(edit_id),)).fetchone()
        if u: form = dict(u)
    return render_template('users.html', users=db.list_all_users(), form=form,
                           edit_id=edit_id, active_page='users')

@app.route('/users/edit/<int:id>')
@login_required
@admin_required
def users_edit(id):
    return redirect(url_for('users', edit_id=id))

@app.route('/users/delete/<int:id>')
@login_required
@admin_required
def users_delete(id):
    if id == session.get('user_id'):
        flash('لا يمكنك حذف حسابك الحالي', 'danger')
    else:
        db.delete_user(id)
        flash('تم الحذف', 'success')
    return redirect(url_for('users'))

@app.route('/alerts')
@login_required
@admin_required
def alerts():
    low = db.get_low_stock_medicines()
    near = db.get_near_expiry(days=30)
    expired = db.get_expired_medicines()
    near_list = []
    for m in near:
        try:
            exp = datetime.strptime(m['expiry_date'], "%Y-%m-%d").date()
            days = (exp - datetime.now().date()).days
        except:
            days = "?"
        d = dict(m)
        d['days_left'] = days
        near_list.append(d)
    return render_template('alerts.html', low_stock=low, near_expiry=near_list,
                           expired=expired, active_page='alerts')

@app.route('/sync/status')
@login_required
@admin_required
def sync_status():
    return jsonify(sync_manager.get_sync_status())

@app.route('/sync/now', methods=['POST'])
@login_required
@admin_required
def sync_now():
    success = sync_manager.sync_all()
    flash('تمت المزامنة بنجاح' if success else 'فشلت المزامنة', 'success' if success else 'danger')
    return redirect(url_for('dashboard'))

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
